import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font    = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill    = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_hdr_fill   = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
sub_hdr_font   = Font(name='Calibri', bold=True, size=11)
normal_font    = Font(name='Calibri', size=11)
bold_font      = Font(name='Calibri', bold=True, size=11)
title_font     = Font(name='Calibri', bold=True, size=14, color='2F5496')
subtitle_font  = Font(name='Calibri', bold=True, size=12, color='2F5496')
explain_font   = Font(name='Calibri', size=10, italic=True, color='595959')
section_font   = Font(name='Calibri', bold=True, size=11, color='C00000')
result_fill    = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
best_fill      = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
worst_fill     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill    = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
orange_fill    = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
blue_fill      = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
purple_fill    = PatternFill(start_color='E2D9F3', end_color='E2D9F3', fill_type='solid')

thin = Side(style='thin')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_top   = Border(left=thin, right=thin, top=Side(style='medium'), bottom=thin)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
wrap_left    = Alignment(horizontal='left',   vertical='top',    wrap_text=True)


def style_cell(ws, row, col, bold=False, fill=None, align='center', num_fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font      = bold_font if bold else normal_font
    cell.alignment = center_align if align == 'center' else left_align
    cell.border    = thin_border
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border


def add_note(ws, row, col, text, merge_end_col=None, row_height=None):
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = explain_font
    cell.alignment = wrap_left
    if row_height:
        ws.row_dimensions[row].height = row_height


def section_title(ws, row, col, text, merge_end_col=None):
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = subtitle_font


def divider(ws, row, col_start, col_end, fill=None):
    """Fill a full-width separator row."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c, value='')
        cell.border = thin_border
        if fill:
            cell.fill = fill


# ============================================================
# DATA
# ============================================================
alternatif = {
    'A1': 'Promosi melalui Media Sosial',
    'A2': 'Diskon atau Promo Bundling',
    'A3': 'Kerja Sama Pesan Antar Online',
    'A4': 'Peningkatan Kualitas Pelayanan',
    'A5': 'Program Kode Referral',
    'A6': 'Program Loyalitas Pelanggan',
    'A7': 'Promosi Marketplace/Aggregator',
    'A8': 'Kolaborasi Influencer Mikro Lokal',
}

kriteria = {
    'C1': {'nama': 'Biaya Promosi',                    'tipe': 'Cost',    'bobot': 0.15},
    'C2': {'nama': 'Jangkauan Pasar',                  'tipe': 'Benefit', 'bobot': 0.20},
    'C3': {'nama': 'Kemudahan Implementasi',           'tipe': 'Benefit', 'bobot': 0.15},
    'C4': {'nama': 'Potensi Peningkatan Penjualan',    'tipe': 'Benefit', 'bobot': 0.20},
    'C5': {'nama': 'Tingkat Risiko',                   'tipe': 'Cost',    'bobot': 0.10},
    'C6': {'nama': 'Keterukuran Hasil Promosi',        'tipe': 'Benefit', 'bobot': 0.10},
    'C7': {'nama': 'Keberlanjutan Strategi',           'tipe': 'Benefit', 'bobot': 0.10},
}

matrix = {
    'A1': [2, 8, 9, 8, 3, 9, 8],
    'A2': [3, 6, 8, 7, 4, 6, 6],
    'A3': [4, 7, 7, 8, 5, 7, 7],
    'A4': [1, 7, 8, 7, 2, 5, 8],
    'A5': [2, 6, 7, 7, 3, 6, 7],
    'A6': [3, 7, 8, 8, 3, 7, 9],
    'A7': [5, 9, 7, 9, 5, 8, 7],
    'A8': [4, 7, 8, 8, 4, 7, 7],
}

krit_keys = list(kriteria.keys())   # C1..C7
alt_keys  = list(alternatif.keys()) # A1..A8
v_default = 0.5
NUM_ALT   = len(alt_keys)    # 8
NUM_KRIT  = len(krit_keys)   # 7
NUM_COLS  = 2 + NUM_KRIT     # kode + nama + 7 criteria = 9 cols


# ================================================================
#  SHEET 1 — Input Data, Matriks Keputusan, f+ dan f−
# ================================================================
ws1 = wb.active
ws1.title = 'Input & Matriks Keputusan'

# ---- HEADER ----
ws1.merge_cells(f'A1:{get_column_letter(NUM_COLS)}1')
ws1.cell(row=1, column=1,
         value='SHEET 1 — INPUT DATA, MATRIKS KEPUTUSAN, DAN NILAI IDEAL (f⁺ / f⁻)').font = title_font
ws1.row_dimensions[1].height = 22

# ================================================================
# BAGIAN A — Daftar Kriteria
# ================================================================
R = 3   # current row pointer
section_title(ws1, R, 1, 'A. DAFTAR KRITERIA DAN BOBOT', merge_end_col=5)
R += 1
add_note(ws1, R, 1,
    'Kriteria terdiri dari 7 aspek. Tipe Cost = semakin kecil semakin baik; '
    'Benefit = semakin besar semakin baik. Total bobot = 1.00.',
    merge_end_col=5, row_height=28)
R += 1

# Header kriteria
hdr_k = ['No', 'Kode', 'Nama Kriteria', 'Tipe', 'Bobot (Wj)']
for i, h in enumerate(hdr_k, 1):
    ws1.cell(row=R, column=i, value=h)
style_header_row(ws1, R, len(hdr_k))
ws1.row_dimensions[R].height = 18
KRIT_HDR = R
R += 1

KRIT_DATA_START = R
for idx, (kode, info) in enumerate(kriteria.items(), 1):
    ws1.cell(row=R, column=1, value=idx)
    ws1.cell(row=R, column=2, value=kode)
    ws1.cell(row=R, column=3, value=info['nama'])
    ws1.cell(row=R, column=4, value=info['tipe'])
    ws1.cell(row=R, column=5, value=info['bobot'])
    for c in range(1, 6):
        fill = blue_fill if idx % 2 == 0 else None
        style_cell(ws1, R, c, fill=fill, num_fmt='0.00' if c == 5 else None)
    R += 1

KRIT_DATA_END = R - 1
# Total bobot
ws1.cell(row=R, column=4, value='TOTAL BOBOT')
ws1.cell(row=R, column=5, value=f'=SUM(E{KRIT_DATA_START}:E{KRIT_DATA_END})')
for c in range(4, 6):
    style_cell(ws1, R, c, bold=True, fill=sub_hdr_fill,
               num_fmt='0.00' if c == 5 else None)
R += 2

# ================================================================
# BAGIAN B — Daftar Alternatif
# ================================================================
section_title(ws1, R, 1, 'B. DAFTAR ALTERNATIF STRATEGI PEMASARAN', merge_end_col=5)
R += 1
add_note(ws1, R, 1,
    '8 alternatif strategi pemasaran yang akan dievaluasi. '
    'Masing-masing dinilai berdasarkan 7 kriteria di atas.',
    merge_end_col=5, row_height=24)
R += 1

hdr_a = ['No', 'Kode', 'Nama Alternatif']
for i, h in enumerate(hdr_a, 1):
    ws1.cell(row=R, column=i, value=h)
style_header_row(ws1, R, len(hdr_a))
ws1.row_dimensions[R].height = 18
R += 1

for idx, (kode, nama) in enumerate(alternatif.items(), 1):
    ws1.cell(row=R, column=1, value=idx)
    ws1.cell(row=R, column=2, value=kode)
    ws1.cell(row=R, column=3, value=nama)
    fill = blue_fill if idx % 2 == 0 else None
    for c in range(1, 4):
        style_cell(ws1, R, c, fill=fill)
    R += 1
R += 1

# ================================================================
# BAGIAN C — Matriks Keputusan
# ================================================================
section_title(ws1, R, 1, 'C. MATRIKS KEPUTUSAN (fij)', merge_end_col=NUM_COLS)
R += 1
add_note(ws1, R, 1,
    'Nilai penilaian setiap alternatif terhadap setiap kriteria (skala 1–9). '
    'Nilai lebih tinggi berarti lebih baik, KECUALI untuk kriteria Cost (semakin rendah = lebih baik).',
    merge_end_col=NUM_COLS, row_height=30)
R += 1

# Baris tipe kriteria
ws1.cell(row=R, column=1, value='Tipe')
ws1.cell(row=R, column=2, value='')
style_cell(ws1, R, 1, bold=True, fill=sub_hdr_fill)
style_cell(ws1, R, 2, bold=True, fill=sub_hdr_fill)
TIPE_ROW = R
for c_idx, k in enumerate(krit_keys, 3):
    ws1.cell(row=R, column=c_idx, value=kriteria[k]['tipe'])
    style_cell(ws1, R, c_idx, bold=True, fill=sub_hdr_fill)
R += 1

# Header matriks
hdr_m = ['Kode', 'Alternatif'] + [f'{k}\n({kriteria[k]["nama"]})' for k in krit_keys]
for i, h in enumerate(hdr_m, 1):
    ws1.cell(row=R, column=i, value=h)
style_header_row(ws1, R, NUM_COLS)
ws1.row_dimensions[R].height = 36
MK_HEADER_ROW = R
R += 1

DATA_START_ROW = R
for idx, (kode, vals) in enumerate(matrix.items(), 1):
    ws1.cell(row=R, column=1, value=kode)
    ws1.cell(row=R, column=2, value=alternatif[kode])
    for c_offset, val in enumerate(vals):
        ws1.cell(row=R, column=3 + c_offset, value=val)
    fill = blue_fill if idx % 2 == 0 else None
    for c in range(1, NUM_COLS + 1):
        style_cell(ws1, R, c, fill=fill)
    R += 1
DATA_END_ROW = R - 1   # row 14 (if data starts at 7 with 8 alts)

# Baris Bobot
BOBOT_ROW_S1 = R
ws1.cell(row=R, column=1, value='')
ws1.cell(row=R, column=2, value='Bobot (Wj)')
BOBOT_COL_START = 3
for c_idx, k in enumerate(krit_keys, BOBOT_COL_START):
    ws1.cell(row=R, column=c_idx, value=kriteria[k]['bobot'])
for c in range(1, NUM_COLS + 1):
    style_cell(ws1, R, c, bold=True, fill=sub_hdr_fill,
               num_fmt='0.00' if c >= BOBOT_COL_START else None)
R += 2

# ================================================================
# BAGIAN D — f+ dan f−
# ================================================================
section_title(ws1, R, 1, 'D. NILAI IDEAL TERBAIK (f⁺) DAN TERBURUK (f⁻)', merge_end_col=NUM_COLS)
R += 1
add_note(ws1, R, 1,
    'f⁺ (Best): untuk Benefit = MAX, untuk Cost = MIN.  '
    'f⁻ (Worst): untuk Benefit = MIN, untuk Cost = MAX.  '
    'Rumus Excel IF digunakan untuk membedakan tipe kriteria secara otomatis.',
    merge_end_col=NUM_COLS, row_height=32)
R += 1

# f+ row
ws1.cell(row=R, column=1, value='')
ws1.cell(row=R, column=2, value='f⁺ (Nilai Terbaik)')
FPLUS_ROW = R
for c_idx, k in enumerate(krit_keys, 3):
    col_letter = get_column_letter(c_idx)
    tipe_ref   = f'{col_letter}{TIPE_ROW}'
    data_range = f'{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}'
    formula    = f'=IF({tipe_ref}="Cost",MIN({data_range}),MAX({data_range}))'
    ws1.cell(row=R, column=c_idx, value=formula)
for c in range(1, NUM_COLS + 1):
    style_cell(ws1, R, c, bold=True, fill=best_fill)
R += 1

# f- row
ws1.cell(row=R, column=1, value='')
ws1.cell(row=R, column=2, value='f⁻ (Nilai Terburuk)')
FMINUS_ROW = R
for c_idx, k in enumerate(krit_keys, 3):
    col_letter = get_column_letter(c_idx)
    tipe_ref   = f'{col_letter}{TIPE_ROW}'
    data_range = f'{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}'
    formula    = f'=IF({tipe_ref}="Cost",MAX({data_range}),MIN({data_range}))'
    ws1.cell(row=R, column=c_idx, value=formula)
for c in range(1, NUM_COLS + 1):
    style_cell(ws1, R, c, bold=True, fill=worst_fill)
R += 2

# Contoh rumus f+/f-
add_note(ws1, R, 1,
    f'📌 Contoh rumus f⁺ di C{FPLUS_ROW} (C1, Cost): '
    f'=IF(C{TIPE_ROW}="Cost",MIN(C{DATA_START_ROW}:C{DATA_END_ROW}),MAX(C{DATA_START_ROW}:C{DATA_END_ROW})) '
    f'→ MIN(2,3,4,1,2,3,5,4) = 1',
    merge_end_col=NUM_COLS, row_height=22)
R += 1
add_note(ws1, R, 1,
    f'📌 Contoh rumus f⁻ di C{FMINUS_ROW} (C1, Cost): '
    f'=IF(C{TIPE_ROW}="Cost",MAX(C{DATA_START_ROW}:C{DATA_END_ROW}),MIN(C{DATA_START_ROW}:C{DATA_END_ROW})) '
    f'→ MAX(2,3,4,1,2,3,5,4) = 5',
    merge_end_col=NUM_COLS, row_height=22)

# Column widths
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 42
for col in range(3, NUM_COLS + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 18
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14

# ================================================================
#  SHEET 2 — Normalisasi, F*ij Terbobot, Perhitungan S dan R
# ================================================================
ws2 = wb.create_sheet('Normalisasi & S, R')

SH1 = "'Input & Matriks Keputusan'"   # reference string to sheet 1

ws2.merge_cells(f'A1:{get_column_letter(NUM_COLS)}1')
ws2.cell(row=1, column=1,
         value='SHEET 2 — NORMALISASI (Nij), F*ij TERBOBOT, DAN PERHITUNGAN Si / Ri').font = title_font
ws2.row_dimensions[1].height = 22

R2 = 3  # row pointer for sheet 2

# ================================================================
# BAGIAN A — Matriks Normalisasi (Nij)
# ================================================================
section_title(ws2, R2, 1, 'A. MATRIKS NORMALISASI (Nij)', merge_end_col=NUM_COLS)
R2 += 1
add_note(ws2, R2, 1,
    'Nij = (f⁺j − fij) / (f⁺j − f⁻j)   →   Nij=0 berarti terbaik, Nij=1 berarti terburuk. '
    'Rumus mereferensikan data dari Sheet 1.',
    merge_end_col=NUM_COLS, row_height=28)
R2 += 1

# Formula label
ws2.merge_cells(f'A{R2}:{get_column_letter(NUM_COLS)}{R2}')
ws2.cell(row=R2, column=1, value='Rumus: Nij = (f⁺j − fij) / (f⁺j − f⁻j)').font = section_font
R2 += 1

# Header
hdr_n = ['Kode', 'Alternatif'] + [f'{k}\nNij' for k in krit_keys]
for i, h in enumerate(hdr_n, 1):
    ws2.cell(row=R2, column=i, value=h)
style_header_row(ws2, R2, NUM_COLS)
ws2.row_dimensions[R2].height = 30
R2 += 1

NIJ_START = R2
for r_offset, kode in enumerate(alt_keys):
    r       = R2 + r_offset
    data_r  = DATA_START_ROW + r_offset   # row in sheet 1

    ws2.cell(row=r, column=1, value=kode)
    ws2.cell(row=r, column=2, value=alternatif[kode])

    for c_off, k in enumerate(krit_keys):
        c          = 3 + c_off
        col_letter = get_column_letter(c)
        fplus_ref  = f"{SH1}!{col_letter}{FPLUS_ROW}"
        fminus_ref = f"{SH1}!{col_letter}{FMINUS_ROW}"
        fij_ref    = f"{SH1}!{col_letter}{data_r}"
        formula    = (f'=IF(({fplus_ref}-{fminus_ref})=0,0,'
                      f'({fplus_ref}-{fij_ref})/({fplus_ref}-{fminus_ref}))')
        ws2.cell(row=r, column=c, value=formula)
        ws2.cell(row=r, column=c).number_format = '0.0000'

    fill = blue_fill if (r_offset + 1) % 2 == 0 else None
    for c in range(1, NUM_COLS + 1):
        style_cell(ws2, r, c, fill=fill)

NIJ_END = NIJ_START + NUM_ALT - 1
R2 = NIJ_END + 2

add_note(ws2, R2, 1,
    f'📌 Contoh N(A1,C1): C1 = Cost → f⁺=1, f⁻=5, f(A1)=2 '
    f'→ N = (1−2)/(1−5) = (−1)/(−4) = 0.2500',
    merge_end_col=NUM_COLS, row_height=22)
R2 += 1
add_note(ws2, R2, 1,
    f'📌 Contoh N(A1,C2): C2 = Benefit → f⁺=9, f⁻=6, f(A1)=8 '
    f'→ N = (9−8)/(9−6) = 1/3 = 0.3333',
    merge_end_col=NUM_COLS, row_height=22)
R2 += 2

# ================================================================
# BAGIAN B — F*ij Terbobot
# ================================================================
section_title(ws2, R2, 1, 'B. MATRIKS NORMALISASI TERBOBOT (F*ij)', merge_end_col=NUM_COLS)
R2 += 1
add_note(ws2, R2, 1,
    'F*ij = Wj × Nij  →  nilai normalisasi dikali bobot kriteria. '
    'Hasilnya digunakan untuk menghitung Si (Utility) dan Ri (Regret).',
    merge_end_col=NUM_COLS, row_height=28)
R2 += 1

ws2.merge_cells(f'A{R2}:{get_column_letter(NUM_COLS)}{R2}')
ws2.cell(row=R2, column=1, value='Rumus: F*ij = Wj × Nij').font = section_font
R2 += 1

# Bobot reference row
ws2.cell(row=R2, column=1, value='')
ws2.cell(row=R2, column=2, value='Bobot (Wj)')
BOBOT_ROW_S2 = R2
for c_idx, k in enumerate(krit_keys, 3):
    col_letter = get_column_letter(c_idx)
    ws2.cell(row=R2, column=c_idx,
             value=f"={SH1}!{col_letter}{BOBOT_ROW_S1}")
    ws2.cell(row=R2, column=c_idx).number_format = '0.00'
for c in range(1, NUM_COLS + 1):
    style_cell(ws2, R2, c, bold=True, fill=sub_hdr_fill)
R2 += 1

# Header
hdr_f = ['Kode', 'Alternatif'] + [f'{k}\nF*ij' for k in krit_keys]
for i, h in enumerate(hdr_f, 1):
    ws2.cell(row=R2, column=i, value=h)
style_header_row(ws2, R2, NUM_COLS)
ws2.row_dimensions[R2].height = 30
R2 += 1

FIJ_START = R2
for r_offset, kode in enumerate(alt_keys):
    r       = R2 + r_offset
    nij_r   = NIJ_START + r_offset

    ws2.cell(row=r, column=1, value=kode)
    ws2.cell(row=r, column=2, value=alternatif[kode])

    for c_off, k in enumerate(krit_keys):
        c          = 3 + c_off
        col_letter = get_column_letter(c)
        bobot_ref  = f"={SH1}!{col_letter}{BOBOT_ROW_S1}*{col_letter}{nij_r}"
        ws2.cell(row=r, column=c, value=bobot_ref)
        ws2.cell(row=r, column=c).number_format = '0.0000'

    fill = orange_fill if (r_offset + 1) % 2 == 0 else None
    for c in range(1, NUM_COLS + 1):
        style_cell(ws2, r, c, fill=fill)

FIJ_END = FIJ_START + NUM_ALT - 1
R2 = FIJ_END + 2

add_note(ws2, R2, 1,
    f'📌 Contoh F*(A1,C1) = W1 × N(A1,C1) = 0.15 × 0.2500 = 0.0375',
    merge_end_col=NUM_COLS, row_height=20)
R2 += 2

# ================================================================
# BAGIAN C — Perhitungan Si dan Ri
# ================================================================
section_title(ws2, R2, 1, 'C. PERHITUNGAN Si (UTILITY MEASURE) DAN Ri (REGRET MEASURE)', merge_end_col=8)
R2 += 1
add_note(ws2, R2, 1,
    'Si = Σ F*ij  →  jumlah total penyimpangan terbobot; semakin kecil = semakin baik secara keseluruhan.\n'
    'Ri = MAX(F*ij)  →  penyimpangan terbesar pada satu kriteria; semakin kecil = tidak ada kriteria "terburuk".',
    merge_end_col=8, row_height=38)
R2 += 1

ws2.merge_cells(f'A{R2}:H{R2}')
ws2.cell(row=R2, column=1,
         value='Rumus Si = SUM(F*ij)   |   Rumus Ri = MAX(F*ij)').font = section_font
R2 += 1

# Header Si/Ri
hdr_sr = ['Kode', 'Alternatif', 'Si (Utility Measure)', 'Ri (Regret Measure)']
for i, h in enumerate(hdr_sr, 1):
    ws2.cell(row=R2, column=i, value=h)
style_header_row(ws2, R2, len(hdr_sr))
ws2.row_dimensions[R2].height = 18
SR_HDR = R2
R2 += 1

SR_START = R2
for r_offset, a in enumerate(alt_keys):
    r     = R2 + r_offset
    fij_r = FIJ_START + r_offset

    ws2.cell(row=r, column=1, value=a)
    ws2.cell(row=r, column=2, value=alternatif[a])

    # Si = SUM across all criteria columns (col 3 to col 9)
    last_krit_col = get_column_letter(2 + NUM_KRIT)
    si_formula = f'=SUM(C{fij_r}:{last_krit_col}{fij_r})'
    ws2.cell(row=r, column=3, value=si_formula)
    ws2.cell(row=r, column=3).number_format = '0.0000'

    ri_formula = f'=MAX(C{fij_r}:{last_krit_col}{fij_r})'
    ws2.cell(row=r, column=4, value=ri_formula)
    ws2.cell(row=r, column=4).number_format = '0.0000'

    fill = result_fill if (r_offset + 1) % 2 == 0 else None
    for c in range(1, 5):
        style_cell(ws2, r, c, fill=fill)

SR_END = SR_START + NUM_ALT - 1
R2 = SR_END + 1

# S−, S+, R−, R+ values
si_range = f'C{SR_START}:C{SR_END}'
ri_range = f'D{SR_START}:D{SR_END}'

SMIN_ROW = R2
ws2.cell(row=R2, column=2, value='S⁻ = MIN(Si) — paling ideal')
ws2.cell(row=R2, column=3, value=f'=MIN({si_range})')
ws2.cell(row=R2, column=3).number_format = '0.0000'
ws2.cell(row=R2, column=4, value='R⁻ = MIN(Ri)')
ws2.cell(row=R2, column=5, value=f'=MIN({ri_range})')
ws2.cell(row=R2, column=5).number_format = '0.0000'
for c in [1, 2, 3, 4, 5]:
    style_cell(ws2, R2, c, bold=True, fill=best_fill)
R2 += 1

SMAX_ROW = R2
ws2.cell(row=R2, column=2, value='S⁺ = MAX(Si) — paling tidak ideal')
ws2.cell(row=R2, column=3, value=f'=MAX({si_range})')
ws2.cell(row=R2, column=3).number_format = '0.0000'
ws2.cell(row=R2, column=4, value='R⁺ = MAX(Ri)')
ws2.cell(row=R2, column=5, value=f'=MAX({ri_range})')
ws2.cell(row=R2, column=5).number_format = '0.0000'
for c in [1, 2, 3, 4, 5]:
    style_cell(ws2, R2, c, bold=True, fill=worst_fill)
R2 += 2

add_note(ws2, R2, 1,
    f'📌 Contoh: Si(A1) = Σ F*(A1,Cj) = 0.0375+0.0667+0.0000+0.1000+0.0333+0.0000+0.0333 = 0.2708\n'
    f'          Ri(A1) = MAX(F*) = 0.1000  (penyimpangan terbesar ada di C4)',
    merge_end_col=8, row_height=38)

# Column widths sheet 2
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 42
for col in range(3, NUM_COLS + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 16

# ================================================================
#  SHEET 3 — Nilai Q, Uji Sensitivitas v, Validasi Kompromi, Hasil
# ================================================================
ws3 = wb.create_sheet('Nilai Q & Hasil')

SH2 = "'Normalisasi & S, R'"   # reference to sheet 2

ws3.merge_cells('A1:J1')
ws3.cell(row=1, column=1,
         value='SHEET 3 — INDEKS VIKOR (Qi), UJI SENSITIVITAS v, VALIDASI KOMPROMI, DAN HASIL AKHIR').font = title_font
ws3.row_dimensions[1].height = 22

R3 = 3

# ================================================================
# BAGIAN A — Nilai Q (Indeks VIKOR)
# ================================================================
section_title(ws3, R3, 1, 'A. PERHITUNGAN INDEKS VIKOR (Qi)', merge_end_col=8)
R3 += 1
add_note(ws3, R3, 1,
    'Qi = v × (Si − S⁻)/(S⁺ − S⁻) + (1−v) × (Ri − R⁻)/(R⁺ − R⁻)\n'
    'Parameter v menentukan bobot antara kepentingan kelompok (Si) vs. individu terburuk (Ri). '
    'v = 0.5 → keseimbangan (by consensus). Qi terkecil = alternatif terbaik.',
    merge_end_col=8, row_height=40)
R3 += 1

ws3.merge_cells(f'A{R3}:H{R3}')
ws3.cell(row=R3, column=1,
         value='Rumus: Qi = v×(Si−S⁻)/(S⁺−S⁻) + (1−v)×(Ri−R⁻)/(R⁺−R⁻)').font = section_font
R3 += 1

# v parameter cell (editable)
ws3.cell(row=R3, column=1, value='Nilai v (dapat diubah)')
ws3.cell(row=R3, column=2, value='→')
ws3.cell(row=R3, column=3, value=v_default)
V_CELL = f'C{R3}'      # absolute reference for v
V_ROW  = R3
style_cell(ws3, R3, 1, bold=True, fill=yellow_fill)
style_cell(ws3, R3, 2, bold=True, fill=yellow_fill)
style_cell(ws3, R3, 3, bold=True, fill=yellow_fill)
R3 += 1

# Header Qi table
hdr_q = ['Kode', 'Alternatif', 'Si', 'Ri', 'Qi', 'Ranking Q', 'Ranking S', 'Ranking R']
for i, h in enumerate(hdr_q, 1):
    ws3.cell(row=R3, column=i, value=h)
style_header_row(ws3, R3, len(hdr_q))
ws3.row_dimensions[R3].height = 18
QI_HDR = R3
R3 += 1

QI_START = R3
# Reference ranges from sheet 2
si_ref_s2 = f"'{ws2.title}'!C{{}}:{{}}"
ri_ref_s2 = f"'{ws2.title}'!D{{}}:{{}}"

si_col_s2 = f"'{ws2.title}'!C{SR_START}:C{SR_END}"
ri_col_s2 = f"'{ws2.title}'!D{SR_START}:D{SR_END}"
s_min_ref  = f"'{ws2.title}'!C{SMIN_ROW}"
s_max_ref  = f"'{ws2.title}'!C{SMAX_ROW}"
r_min_ref  = f"'{ws2.title}'!E{SMIN_ROW}"
r_max_ref  = f"'{ws2.title}'!E{SMAX_ROW}"

qi_range_abs = f'$E${QI_START}:$E${QI_START + NUM_ALT - 1}'
si_range_abs = f'$C${QI_START}:$C${QI_START + NUM_ALT - 1}'
ri_range_abs = f'$D${QI_START}:$D${QI_START + NUM_ALT - 1}'

for r_offset, a in enumerate(alt_keys):
    r    = R3 + r_offset
    sr_r = SR_START + r_offset

    ws3.cell(row=r, column=1, value=a)
    ws3.cell(row=r, column=2, value=alternatif[a])

    # Si reference from sheet 2
    ws3.cell(row=r, column=3, value=f"='{ws2.title}'!C{sr_r}")
    ws3.cell(row=r, column=3).number_format = '0.0000'

    # Ri reference from sheet 2
    ws3.cell(row=r, column=4, value=f"='{ws2.title}'!D{sr_r}")
    ws3.cell(row=r, column=4).number_format = '0.0000'

    # Qi formula
    qi_formula = (
        f'=IF(({s_max_ref}-{s_min_ref})=0,0,'
        f'${V_CELL}*(C{r}-{s_min_ref})/({s_max_ref}-{s_min_ref}))'
        f'+IF(({r_max_ref}-{r_min_ref})=0,0,'
        f'(1-${V_CELL})*(D{r}-{r_min_ref})/({r_max_ref}-{r_min_ref}))'
    )
    ws3.cell(row=r, column=5, value=qi_formula)
    ws3.cell(row=r, column=5).number_format = '0.0000'

    # Ranking Q (ascending — smallest Qi = rank 1)
    ws3.cell(row=r, column=6, value=f'=RANK(E{r},{qi_range_abs},1)')

    # Ranking Si
    ws3.cell(row=r, column=7, value=f'=RANK(C{r},{si_range_abs},1)')

    # Ranking Ri
    ws3.cell(row=r, column=8, value=f'=RANK(D{r},{ri_range_abs},1)')

    fill = result_fill if (r_offset + 1) % 2 == 0 else None
    for c in range(1, 9):
        style_cell(ws3, r, c, fill=fill)

QI_END = QI_START + NUM_ALT - 1
R3 = QI_END + 2

add_note(ws3, R3, 1,
    f'📌 Contoh Qi(A1): v=0.5, Si(A1)=0.2708=S⁻, Ri(A1)=0.1000=R⁻ → Qi = 0.5×0/0.5209 + 0.5×0/0.1000 = 0.0000\n'
    f'   Qi=0 karena A1 memiliki nilai Si dan Ri terkecil → ALTERNATIF TERBAIK.',
    merge_end_col=8, row_height=38)
R3 += 2

# ================================================================
# BAGIAN B — Uji Sensitivitas v
# ================================================================
section_title(ws3, R3, 1, 'B. UJI SENSITIVITAS PARAMETER v', merge_end_col=8)
R3 += 1
add_note(ws3, R3, 1,
    'Uji sensitivitas mengevaluasi apakah ranking berubah jika nilai v berubah. '
    'v=0 → hanya memperhitungkan Ri (regret minimum), v=1 → hanya memperhitungkan Si (utility). '
    'Solusi kompromi yang robust tidak bergantung pada pilihan v.',
    merge_end_col=8, row_height=36)
R3 += 1

v_test_values = [0.0, 0.25, 0.5, 0.75, 1.0]

# Header sensitivitas
hdr_sens = ['v'] + [f'{a}' for a in alt_keys] + ['Rank 1']
for i, h in enumerate(hdr_sens, 1):
    ws3.cell(row=R3, column=i, value=h)
style_header_row(ws3, R3, len(hdr_sens))
ws3.row_dimensions[R3].height = 18
R3 += 1

SENS_START = R3
for v_val in v_test_values:
    ws3.cell(row=R3, column=1, value=v_val)
    style_cell(ws3, R3, 1, bold=True, fill=yellow_fill)

    qi_vals = []
    for r_offset, a in enumerate(alt_keys):
        sr_r   = SR_START + r_offset
        si_ref = f"'{ws2.title}'!C{sr_r}"
        ri_ref = f"'{ws2.title}'!D{sr_r}"
        qi_v   = (
            f'=IF(({s_max_ref}-{s_min_ref})=0,0,'
            f'{v_val}*({si_ref}-{s_min_ref})/({s_max_ref}-{s_min_ref}))'
            f'+IF(({r_max_ref}-{r_min_ref})=0,0,'
            f'{1-v_val}*({ri_ref}-{r_min_ref})/({r_max_ref}-{r_min_ref}))'
        )
        c = 2 + r_offset
        ws3.cell(row=R3, column=c, value=qi_v)
        ws3.cell(row=R3, column=c).number_format = '0.0000'
        qi_vals.append(f'{get_column_letter(c)}{R3}')
        fill = blue_fill if (r_offset + 1) % 2 == 0 else None
        style_cell(ws3, R3, c, fill=fill)

    # Rank 1 = alternative with min Qi
    qi_range_sens = ','.join(qi_vals)
    rank1_col = 2 + NUM_ALT
    alt_range_abs_qi = ','.join([f'{get_column_letter(2+i)}{R3}' for i in range(NUM_ALT)])
    # Use INDEX+MATCH to find which alt has min value
    qi_cells_range = f'{get_column_letter(2)}{R3}:{get_column_letter(1+NUM_ALT)}{R3}'
    alt_labels      = list(alt_keys)
    # write the alt_labels as helper — just use MATCH+INDEX on the row
    formula_rank1 = (
        f'=INDEX({{"' + '","'.join(alt_keys) + f'"}},MATCH(MIN({qi_cells_range}),{qi_cells_range},0))'
    )
    ws3.cell(row=R3, column=rank1_col, value=formula_rank1)
    style_cell(ws3, R3, rank1_col, bold=True, fill=best_fill)

    R3 += 1

SENS_END = R3 - 1
R3 += 1

add_note(ws3, R3, 1,
    '📌 Interpretasi: Jika alternatif rank 1 konsisten (sama) untuk semua nilai v, '
    'maka solusi kompromi bersifat ROBUST (tidak sensitif terhadap pemilihan v).',
    merge_end_col=8, row_height=28)
R3 += 2

# ================================================================
# BAGIAN C — Validasi Solusi Kompromi
# ================================================================
section_title(ws3, R3, 1, 'C. VALIDASI SOLUSI KOMPROMI', merge_end_col=6)
R3 += 1
add_note(ws3, R3, 1,
    'Solusi kompromi (alternatif rank-1 pada Qi) divalidasi melalui DUA kondisi:\n'
    '  Kondisi 1 — Acceptable Advantage: Q(rank2) − Q(rank1) ≥ DQ = 1/(m−1)\n'
    '  Kondisi 2 — Acceptable Stability: rank-1 pada Q juga rank-1 pada Si ATAU Ri\n'
    'Jika kedua kondisi terpenuhi → solusi kompromi SAH.',
    merge_end_col=6, row_height=52)
R3 += 2

# --- Kondisi 1 ---
section_title(ws3, R3, 1, 'KONDISI 1: Acceptable Advantage', merge_end_col=4)
R3 += 1

qi_col_abs = f'$E${QI_START}:$E${QI_END}'

ws3.cell(row=R3, column=1, value='Jumlah alternatif (m)')
ws3.cell(row=R3, column=2, value=NUM_ALT)
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
M_CELL = f'B{R3}'
R3 += 1

ws3.cell(row=R3, column=1, value='DQ = 1/(m−1)')
ws3.cell(row=R3, column=2, value=f'=1/({M_CELL}-1)')
ws3.cell(row=R3, column=2).number_format = '0.0000'
DQ_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, fill=yellow_fill)
R3 += 1

ws3.cell(row=R3, column=1, value='Q Peringkat 1 (Q₁)')
ws3.cell(row=R3, column=2, value=f'=SMALL({qi_col_abs},1)')
ws3.cell(row=R3, column=2).number_format = '0.0000'
Q1_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, fill=best_fill)
R3 += 1

ws3.cell(row=R3, column=1, value='Q Peringkat 2 (Q₂)')
ws3.cell(row=R3, column=2, value=f'=SMALL({qi_col_abs},2)')
ws3.cell(row=R3, column=2).number_format = '0.0000'
Q2_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
R3 += 1

ws3.cell(row=R3, column=1, value='Selisih Q₂ − Q₁')
ws3.cell(row=R3, column=2, value=f'={Q2_CELL}-{Q1_CELL}')
ws3.cell(row=R3, column=2).number_format = '0.0000'
QDIFF_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, fill=yellow_fill)
R3 += 1

ws3.cell(row=R3, column=1, value='Kondisi 1 terpenuhi?')
ws3.cell(row=R3, column=2,
         value=f'=IF({QDIFF_CELL}>={DQ_CELL},"✓ TERPENUHI","✗ TIDAK TERPENUHI")')
KOND1_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, bold=True)
R3 += 2

# --- Kondisi 2 ---
section_title(ws3, R3, 1, 'KONDISI 2: Acceptable Stability in Decision Making', merge_end_col=4)
R3 += 1

alt_range_q = f'$A${QI_START}:$A${QI_END}'
si_col_here  = f'$C${QI_START}:$C${QI_END}'
ri_col_here  = f'$D${QI_START}:$D${QI_END}'

ws3.cell(row=R3, column=1, value='Alternatif Rank 1 (Q terkecil)')
ws3.cell(row=R3, column=2,
         value=f'=INDEX({alt_range_q},MATCH(SMALL({qi_col_abs},1),{qi_col_abs},0))')
ALT_RANK1_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, bold=True, fill=best_fill)
R3 += 1

ws3.cell(row=R3, column=1, value='Si alternatif Rank 1')
ws3.cell(row=R3, column=2,
         value=f'=INDEX({si_col_here},MATCH(SMALL({qi_col_abs},1),{qi_col_abs},0))')
ws3.cell(row=R3, column=2).number_format = '0.0000'
SI_R1_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
R3 += 1

ws3.cell(row=R3, column=1, value='Ranking Si alternatif Rank 1')
ws3.cell(row=R3, column=2, value=f'=RANK({SI_R1_CELL},{si_col_here},1)')
RANK_SI_R1 = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
R3 += 1

ws3.cell(row=R3, column=1, value='Ri alternatif Rank 1')
ws3.cell(row=R3, column=2,
         value=f'=INDEX({ri_col_here},MATCH(SMALL({qi_col_abs},1),{qi_col_abs},0))')
ws3.cell(row=R3, column=2).number_format = '0.0000'
RI_R1_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
R3 += 1

ws3.cell(row=R3, column=1, value='Ranking Ri alternatif Rank 1')
ws3.cell(row=R3, column=2, value=f'=RANK({RI_R1_CELL},{ri_col_here},1)')
RANK_RI_R1 = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2)
R3 += 1

ws3.cell(row=R3, column=1, value='Kondisi 2 terpenuhi?')
ws3.cell(row=R3, column=2,
         value=f'=IF(OR({RANK_SI_R1}=1,{RANK_RI_R1}=1),"✓ TERPENUHI","✗ TIDAK TERPENUHI")')
KOND2_CELL = f'B{R3}'
style_cell(ws3, R3, 1, bold=True)
style_cell(ws3, R3, 2, bold=True)
R3 += 2

# ================================================================
# BAGIAN D — Hasil Akhir / Ranking Final
# ================================================================
section_title(ws3, R3, 1, 'D. HASIL AKHIR DAN REKOMENDASI', merge_end_col=8)
R3 += 1

# Kesimpulan validasi
ws3.merge_cells(f'A{R3}:H{R3}')
ws3.cell(row=R3, column=1,
    value=(
        f'=IF(AND({KOND1_CELL}="✓ TERPENUHI",{KOND2_CELL}="✓ TERPENUHI"),'
        f'"✓ Kedua kondisi TERPENUHI → "&{ALT_RANK1_CELL}&" adalah SOLUSI KOMPROMI TERBAIK yang SAH",'
        f'"✗ Salah satu kondisi TIDAK TERPENUHI → Perlu analisis tambahan atau pertimbangkan set solusi")'
    ))
ws3.cell(row=R3, column=1).font = Font(name='Calibri', bold=True, size=13, color='2F5496')
ws3.cell(row=R3, column=1).alignment = center_align
ws3.cell(row=R3, column=1).border    = thin_border
ws3.cell(row=R3, column=1).fill     = result_fill
ws3.row_dimensions[R3].height = 26
R3 += 2

# Ranking tabel final
hdr_final = ['Rank', 'Kode', 'Nama Alternatif', 'Si', 'Ri', 'Qi']
for i, h in enumerate(hdr_final, 1):
    ws3.cell(row=R3, column=i, value=h)
style_header_row(ws3, R3, len(hdr_final))
ws3.row_dimensions[R3].height = 18
R3 += 1

FINAL_START = R3
for rank in range(1, NUM_ALT + 1):
    # Use INDEX/MATCH to find the alternative with this rank
    formula_kode = (
        f'=INDEX($A${QI_START}:$A${QI_END},'
        f'MATCH(SMALL($E${QI_START}:$E${QI_END},{rank}),'
        f'$E${QI_START}:$E${QI_END},0))'
    )
    formula_nama = (
        f'=INDEX($B${QI_START}:$B${QI_END},'
        f'MATCH(SMALL($E${QI_START}:$E${QI_END},{rank}),'
        f'$E${QI_START}:$E${QI_END},0))'
    )
    formula_si = (
        f'=SMALL($C${QI_START}:$C${QI_END},{rank})'
    )
    formula_ri = (
        f'=INDEX($D${QI_START}:$D${QI_END},'
        f'MATCH(SMALL($E${QI_START}:$E${QI_END},{rank}),'
        f'$E${QI_START}:$E${QI_END},0))'
    )
    formula_qi = f'=SMALL($E${QI_START}:$E${QI_END},{rank})'

    ws3.cell(row=R3, column=1, value=rank)
    ws3.cell(row=R3, column=2, value=formula_kode)
    ws3.cell(row=R3, column=3, value=formula_nama)
    ws3.cell(row=R3, column=4, value=formula_si)
    ws3.cell(row=R3, column=5, value=formula_ri)
    ws3.cell(row=R3, column=6, value=formula_qi)

    ws3.cell(row=R3, column=4).number_format = '0.0000'
    ws3.cell(row=R3, column=5).number_format = '0.0000'
    ws3.cell(row=R3, column=6).number_format = '0.0000'

    if rank == 1:
        fill = best_fill
        bold = True
    elif rank == NUM_ALT:
        fill = worst_fill
        bold = False
    else:
        fill = result_fill if rank % 2 == 0 else None
        bold = False

    for c in range(1, 7):
        style_cell(ws3, R3, c, bold=bold, fill=fill)

    R3 += 1

FINAL_END = R3 - 1
R3 += 2

add_note(ws3, R3, 1,
    '📌 Tabel Hasil Akhir mengurutkan alternatif dari terbaik (Qi terkecil) ke terburuk. '
    'Kolom Si, Ri, Qi diurutkan berdasarkan ranking Qi menggunakan INDEX/MATCH/SMALL. '
    'Apabila data input atau bobot diubah di Sheet 1, seluruh hasil akan otomatis ter-update.',
    merge_end_col=8, row_height=38)

# Column widths sheet 3
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 42
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 16
ws3.column_dimensions['G'].width = 16
ws3.column_dimensions['H'].width = 16
ws3.column_dimensions['I'].width = 16
ws3.column_dimensions['J'].width = 16

# ============================================================
# SAVE
# ============================================================
filepath = r'c:\kuliah\spk\Perhitungan_Manual_VIKOR_v2.xlsx'
wb.save(filepath)
print(f'[OK] Excel berhasil disimpan: {filepath}')
print(f'\nFile berisi {len(wb.sheetnames)} sheet:')
for i, name in enumerate(wb.sheetnames, 1):
    print(f'  {i}. {name}')
print('\nStruktur sheet:')
print('  Sheet 1 - Input & Matriks Keputusan : Input data, matriks fij, f+ dan f-')
print('  Sheet 2 - Normalisasi & S, R        : Nij, F*ij terbobot, Si dan Ri')
print('  Sheet 3 - Nilai Q & Hasil           : Qi, uji sensitivitas v, validasi kompromi, hasil akhir')
